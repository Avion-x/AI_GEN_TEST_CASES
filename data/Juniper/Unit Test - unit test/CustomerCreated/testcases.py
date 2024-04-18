import openpyxl
import openai
import re

openai.api_key = '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'

def generate_test_scripts(prompt):
    try:
        response = openai.ChatCompletion.create(
            model='gpt-3.5-turbo',  
            messages=[
                {"role": "system", "content": "You are a helpful assistant that generates test cases for networking equipment."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.5,
            max_tokens=4096  
        )
        if response and response.choices:
            
            if isinstance(response.choices, list):
                return response.choices[0]['message']['content'].strip()
            elif isinstance(response.choices, dict):
                return response.choices['text'].strip()
            else:
                return None
        else:
            return None
    except openai.error.OpenAIError as e:
        print(f"OpenAI API Error: {e}")
        return None


try:
    workbook = openpyxl.load_workbook(r'C:\Users\rspre\Genai_poc\New folder\test_types2.xlsx')
except FileNotFoundError:
    print("File not found. Please check the file path.")
    exit(1)

sheet = workbook.active

result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active
result_sheet.append(["ID", "Test Type", "Test Subcategory", "Heading", "Python Script"])

for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
    test_type = row[0]
    test_sub_category = row[1]
    test_type_id = row[2]

    prompt = f'generate five test case headings mentioned with Test Case to validate the functionality of the {test_sub_category} feature on a router. For each test case, include a clear description of the objective and steps involved.'

    response_headings = generate_test_scripts(prompt)

    headings = re.findall(r'Test Case \d+: (.*?)\n', response_headings)
    print(headings)

    for heading in headings[:5]:
        result_sheet.append([test_type_id, test_type, test_sub_category, heading, ""])
        python_script_prompt = f'Generate all Python test scripts for {heading}'
        python_script = generate_test_scripts(python_script_prompt)
        if python_script:
            result_sheet.append(["", "", "", "", python_script])

result_workbook.save("result_python_scripts4.xlsx")
